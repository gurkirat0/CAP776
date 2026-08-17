from openpyxl import load_workbook
from datetime import datetime

file = "data/12614066.xlsx"

# Open existing workbook
wb = load_workbook(file)

# Open Daily Log sheet
ws = wb["Daily Log"]


# Find next empty row
row = 7

while ws.cell(row=row, column=1).value is not None:
    row += 1

print(f"\nAdding today's data to row {row}")


# Date
date_input = input("Date (YYYY-MM-DD): ")

try:
    date = datetime.strptime(date_input, "%Y-%m-%d")
except ValueError:
    print("❌ Invalid date format.")
    print("Use YYYY-MM-DD, for example 2026-08-18")
    exit()


# Daily information
sleep = int(input("Sleep (min): "))
fitness = int(input("Fitness (min): "))
study = int(input("Study (min): "))
coding = int(input("Coding (min): "))
class_time = int(input("Class (min): "))
classes_attended = int(input("Classes Attended: "))
other = int(input("Other Activities (min): "))


# Feeling
print("\nDay's Feeling:")
print("1. Good")
print("2. Neutral")
print("3. Bad")

choice = input("Choose 1-3: ")

feelings = {
    "1": "Good",
    "2": "Neutral",
    "3": "Bad"
}

if choice not in feelings:
    print("❌ Invalid choice.")
    exit()

feeling = feelings[choice]


# Satisfaction
print("\nSatisfaction Level:")
print("1. High")
print("2. Neutral")
print("3. Low")

choice = input("Choose 1-3: ")

satisfaction_levels = {
    "1": "High",
    "2": "Neutral",
    "3": "Low"
}

if choice not in satisfaction_levels:
    print("❌ Invalid choice.")
    exit()

satisfaction = satisfaction_levels[choice]


# Energy
print("\nEnergy Level:")
print("1. High")
print("2. Neutral")
print("3. Low")

choice = input("Choose 1-3: ")

energy_levels = {
    "1": "High",
    "2": "Neutral",
    "3": "Low"
}

if choice not in energy_levels:
    print("❌ Invalid choice.")
    exit()

energy = energy_levels[choice]


# Notes
notes = input("Notes: ")


# Calculate totals
total_tracked = (
    sleep
    + fitness
    + study
    + coding
    + class_time
    + other
)

free_time = 1440 - total_tracked


# Write data to the next row
ws.cell(row, 1).value = date
ws.cell(row, 2).value = sleep
ws.cell(row, 3).value = fitness
ws.cell(row, 4).value = study
ws.cell(row, 5).value = coding
ws.cell(row, 6).value = class_time
ws.cell(row, 7).value = classes_attended
ws.cell(row, 8).value = other
ws.cell(row, 9).value = total_tracked
ws.cell(row, 10).value = free_time
ws.cell(row, 11).value = feeling
ws.cell(row, 12).value = satisfaction
ws.cell(row, 13).value = energy
ws.cell(row, 14).value = notes

# Date display
ws.cell(row, 1).number_format = "yyyy-mm-dd"


# Save without destroying formatting
wb.save(file)

print("\n✅ Daily data added successfully!")
print(f"Total tracked: {total_tracked} minutes")
print(f"Free/unaccounted: {free_time} minutes")
print(f"Saved in row {row}")